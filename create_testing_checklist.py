#!/usr/bin/env python3
"""
AI SchoolOS Testing Checklist Generator
Creates a comprehensive Excel file with detailed test cases, credentials, and dashboard features.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os

def create_testing_checklist():
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    credentials_sheet = wb.create_sheet("Credentials & Quick Reference")
    test_cases_sheet = wb.create_sheet("Detailed Test Cases")
    issues_sheet = wb.create_sheet("Issues Tracker")
    dashboard_sheet = wb.create_sheet("Test Progress Dashboard")
    reminders_sheet = wb.create_sheet("Automated Reminders")
    
    # Define colors
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    purple_fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Credentials & Quick Reference
    setup_credentials_sheet(credentials_sheet, green_fill, thin_border)
    
    # Sheet 2: Detailed Test Cases
    setup_test_cases_sheet(test_cases_sheet, green_fill, red_fill, yellow_fill, thin_border)
    
    # Sheet 3: Issues Tracker
    setup_issues_sheet(issues_sheet, thin_border)
    
    # Sheet 4: Test Progress Dashboard
    setup_dashboard_sheet(dashboard_sheet, green_fill, red_fill, yellow_fill, thin_border)
    
    # Sheet 5: Automated Reminders
    setup_reminders_sheet(reminders_sheet, thin_border)
    
    # Save the workbook
    filename = "AI_SchoolOS_Testing_Checklist.xlsx"
    wb.save(filename)
    print(f"✅ Excel file created successfully: {filename}")
    return filename

def setup_credentials_sheet(sheet, green_fill, border):
    """Setup the credentials and quick reference sheet"""
    
    # Title
    sheet['A1'] = "AI SchoolOS - Testing Credentials & Quick Reference"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet.merge_cells('A1:E1')
    
    # Login Credentials
    sheet['A3'] = "Login Credentials"
    sheet['A3'].font = Font(size=14, bold=True)
    
    # Headers
    headers = ['Role', 'Email', 'Password', 'Portal URL', 'Theme Color']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = green_fill
        cell.border = border
    
    # Data
    credentials_data = [
        ['Admin', 'admin@schoolos.com', 'password', '/admin', 'Blue'],
        ['Teacher', 'teacher@schoolos.com', 'password', '/teacher', 'Green'],
        ['Student', 'student@schoolos.com', 'password', '/student', 'Purple'],
        ['Parent', 'parent@schoolos.com', 'password', '/parent', 'Orange'],
        ['Super Admin', 'superadmin@schoolos.com', 'password', '/super-admin', 'Red']
    ]
    
    for row, data in enumerate(credentials_data, 5):
        for col, value in enumerate(data, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = value
            cell.border = border
    
    # Quick Access Links
    sheet['A11'] = "Quick Access Links"
    sheet['A11'].font = Font(size=14, bold=True)
    
    links_data = [
        ['Frontend', 'http://localhost:3000'],
        ['API Gateway', 'http://localhost:8000'],
        ['API Documentation', 'http://localhost:8000/docs'],
        ['Health Check', 'http://localhost:8000/services/health']
    ]
    
    for row, (name, url) in enumerate(links_data, 12):
        sheet.cell(row=row, column=1, value=name).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=url)
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

def setup_test_cases_sheet(sheet, green_fill, red_fill, yellow_fill, border):
    """Setup the detailed test cases sheet"""
    
    # Title
    sheet['A1'] = "AI SchoolOS - Detailed Test Cases"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet.merge_cells('A1:F1')
    
    # Headers
    headers = ['Test ID', 'Test Case', 'Steps', 'Expected Result', 'Status', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = green_fill
        cell.border = border
    
    # Test cases data
    test_cases = [
        # Landing Page Tests
        ['LP-001', 'Page Load', '1. Navigate to http://localhost:3000', 'Page loads without errors', '⬜', ''],
        ['LP-002', 'Service Cards', '1. Check all 12 service cards', 'All cards visible with icons', '⬜', ''],
        ['LP-003', 'Quick Actions', '1. Click each portal button', 'Navigates to correct portal', '⬜', ''],
        ['LP-004', 'Navigation Tabs', '1. Click Overview/Services tabs', 'Content switches appropriately', '⬜', ''],
        ['LP-005', 'Login Button', '1. Click "Login" button', 'Redirects to /login', '⬜', ''],
        ['LP-006', 'Responsive Design', '1. Resize browser window', 'Layout adapts properly', '⬜', ''],
        
        # Login Page Tests
        ['LOGIN-001', 'Page Load', '1. Navigate to /login', 'Login form displays', '⬜', ''],
        ['LOGIN-002', 'Email Validation', '1. Enter invalid email', 'Shows validation error', '⬜', ''],
        ['LOGIN-003', 'Password Toggle', '1. Click show/hide password', 'Toggle works correctly', '⬜', ''],
        ['LOGIN-004', 'Admin Quick Login', '1. Click "Admin" button', 'Fills credentials, redirects to /admin', '⬜', ''],
        ['LOGIN-005', 'Teacher Quick Login', '1. Click "Teacher" button', 'Fills credentials, redirects to /teacher', '⬜', ''],
        ['LOGIN-006', 'Student Quick Login', '1. Click "Student" button', 'Fills credentials, redirects to /student', '⬜', ''],
        ['LOGIN-007', 'Parent Quick Login', '1. Click "Parent" button', 'Fills credentials, redirects to /parent', '⬜', ''],
        ['LOGIN-008', 'Super Admin Quick Login', '1. Click "Super Admin" button', 'Fills credentials, redirects to /super-admin', '⬜', ''],
        
        # Admin Portal Tests
        ['ADMIN-001', 'Access Control', '1. Try /admin without login', 'Shows "Not Authorized"', '⬜', ''],
        ['ADMIN-002', 'Login Access', '1. Login as admin, go to /admin', 'Loads admin dashboard', '⬜', ''],
        ['ADMIN-003', 'Blue Theme', '1. Check sidebar/header colors', 'Blue theme applied', '⬜', ''],
        ['ADMIN-004', 'Dashboard Stats', '1. Check stats cards', 'All stats display correctly', '⬜', ''],
        ['ADMIN-005', 'Navigation Tabs', '1. Click each tab', 'Content switches appropriately', '⬜', ''],
        ['ADMIN-006', 'Quick Actions', '1. Check action buttons', 'All buttons present', '⬜', ''],
        ['ADMIN-007', 'User Info', '1. Check header user info', 'Displays user name correctly', '⬜', ''],
        ['ADMIN-008', 'Logout Function', '1. Click logout', 'Clears session, redirects to login', '⬜', ''],
        ['ADMIN-009', 'Responsive Design', '1. Test on mobile/tablet/desktop', 'Layout adapts properly', '⬜', ''],
        
        # Teacher Portal Tests
        ['TEACHER-001', 'Access Control', '1. Try /teacher without login', 'Shows "Not Authorized"', '⬜', ''],
        ['TEACHER-002', 'Login Access', '1. Login as teacher, go to /teacher', 'Loads teacher dashboard', '⬜', ''],
        ['TEACHER-003', 'Green Theme', '1. Check sidebar/header colors', 'Green theme applied', '⬜', ''],
        ['TEACHER-004', 'Dashboard Stats', '1. Check stats cards', 'All stats display correctly', '⬜', ''],
        ['TEACHER-005', 'Navigation Tabs', '1. Click each tab', 'Content switches appropriately', '⬜', ''],
        ['TEACHER-006', 'Quick Actions', '1. Check action buttons', 'All buttons present', '⬜', ''],
        ['TEACHER-007', 'User Info', '1. Check header user info', 'Displays user name correctly', '⬜', ''],
        ['TEACHER-008', 'Logout Function', '1. Click logout', 'Clears session, redirects to login', '⬜', ''],
        ['TEACHER-009', 'Responsive Design', '1. Test on mobile/tablet/desktop', 'Layout adapts properly', '⬜', ''],
        ['TEACHER-010', 'Role Indicator', '1. Check role indicator', 'Shows "T" in green circle', '⬜', ''],
        
        # Student Portal Tests
        ['STUDENT-001', 'Access Control', '1. Try /student without login', 'Shows "Not Authorized"', '⬜', ''],
        ['STUDENT-002', 'Login Access', '1. Login as student, go to /student', 'Loads student dashboard', '⬜', ''],
        ['STUDENT-003', 'Purple Theme', '1. Check sidebar/header colors', 'Purple theme applied', '⬜', ''],
        ['STUDENT-004', 'Dashboard Stats', '1. Check stats cards', 'All stats display correctly', '⬜', ''],
        ['STUDENT-005', 'Navigation Tabs', '1. Click each tab', 'Content switches appropriately', '⬜', ''],
        ['STUDENT-006', 'Quick Actions', '1. Check action buttons', 'All buttons present', '⬜', ''],
        ['STUDENT-007', 'User Info', '1. Check header user info', 'Displays user name correctly', '⬜', ''],
        ['STUDENT-008', 'Logout Function', '1. Click logout', 'Clears session, redirects to login', '⬜', ''],
        ['STUDENT-009', 'Responsive Design', '1. Test on mobile/tablet/desktop', 'Layout adapts properly', '⬜', ''],
        ['STUDENT-010', 'Role Indicator', '1. Check role indicator', 'Shows "S" in purple circle', '⬜', ''],
        
        # Parent Portal Tests
        ['PARENT-001', 'Access Control', '1. Try /parent without login', 'Shows "Not Authorized"', '⬜', ''],
        ['PARENT-002', 'Login Access', '1. Login as parent, go to /parent', 'Loads parent dashboard', '⬜', ''],
        ['PARENT-003', 'Orange Theme', '1. Check sidebar/header colors', 'Orange theme applied', '⬜', ''],
        ['PARENT-004', 'Dashboard Stats', '1. Check stats cards', 'All stats display correctly', '⬜', ''],
        ['PARENT-005', 'Navigation Tabs', '1. Click each tab', 'Content switches appropriately', '⬜', ''],
        ['PARENT-006', 'Quick Actions', '1. Check action buttons', 'All buttons present', '⬜', ''],
        ['PARENT-007', 'User Info', '1. Check header user info', 'Displays user name correctly', '⬜', ''],
        ['PARENT-008', 'Logout Function', '1. Click logout', 'Clears session, redirects to login', '⬜', ''],
        ['PARENT-009', 'Responsive Design', '1. Test on mobile/tablet/desktop', 'Layout adapts properly', '⬜', ''],
        ['PARENT-010', 'Role Indicator', '1. Check role indicator', 'Shows "P" in orange circle', '⬜', ''],
        
        # Super Admin Portal Tests
        ['SUPER-001', 'Access Control', '1. Try /super-admin without login', 'Shows "Not Authorized"', '⬜', ''],
        ['SUPER-002', 'Login Access', '1. Login as super admin, go to /super-admin', 'Loads super admin dashboard', '⬜', ''],
        ['SUPER-003', 'Red Theme', '1. Check sidebar/header colors', 'Red theme applied', '⬜', ''],
        ['SUPER-004', 'Dashboard Stats', '1. Check stats cards', 'All stats display correctly', '⬜', ''],
        ['SUPER-005', 'Navigation Tabs', '1. Click each tab', 'Content switches appropriately', '⬜', ''],
        ['SUPER-006', 'Quick Actions', '1. Check action buttons', 'All buttons present', '⬜', ''],
        ['SUPER-007', 'User Info', '1. Check header user info', 'Displays user name correctly', '⬜', ''],
        ['SUPER-008', 'Logout Function', '1. Click logout', 'Clears session, redirects to login', '⬜', ''],
        ['SUPER-009', 'Responsive Design', '1. Test on mobile/tablet/desktop', 'Layout adapts properly', '⬜', ''],
        ['SUPER-010', 'Role Indicator', '1. Check role indicator', 'Shows "S" in red circle', '⬜', ''],
        
        # Cross-Portal Access Tests
        ['CROSS-001', 'Admin Access Other Portals', '1. Login as admin, try other portals', 'Shows "Not Authorized" for all', '⬜', ''],
        ['CROSS-002', 'Teacher Access Other Portals', '1. Login as teacher, try other portals', 'Shows "Not Authorized" for all', '⬜', ''],
        ['CROSS-003', 'Student Access Other Portals', '1. Login as student, try other portals', 'Shows "Not Authorized" for all', '⬜', ''],
        ['CROSS-004', 'Parent Access Other Portals', '1. Login as parent, try other portals', 'Shows "Not Authorized" for all', '⬜', ''],
        ['CROSS-005', 'Super Admin Access Other Portals', '1. Login as super admin, try other portals', 'Shows "Not Authorized" for all', '⬜', ''],
        
        # Backend Health Tests
        ['BACKEND-001', 'API Gateway Health', '1. Visit health endpoint', 'Returns health status', '⬜', ''],
        ['BACKEND-002', 'API Documentation', '1. Visit /docs', 'Swagger UI loads', '⬜', ''],
        ['BACKEND-003', 'Service Connectivity', '1. Check all 18 services', 'All services healthy', '⬜', '']
    ]
    
    # Add test cases to sheet
    for row, test_case in enumerate(test_cases, 4):
        for col, value in enumerate(test_case, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = value
            cell.border = border
    
    # Add data validation for Status column
    dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked"', allow_blank=True)
    sheet.add_data_validation(dv)
    dv.add('E4:E100')  # Apply to Status column
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

def setup_issues_sheet(sheet, border):
    """Setup the issues tracker sheet"""
    
    # Title
    sheet['A1'] = "AI SchoolOS - Issues Tracker"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet.merge_cells('A1:I1')
    
    # Headers
    headers = ['Issue ID', 'Portal', 'Issue Description', 'Steps to Reproduce', 'Severity', 'Status', 'Assigned To', 'Date Reported', 'Screenshot']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        cell.border = border
    
    # Add data validation
    severity_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    sheet.add_data_validation(severity_dv)
    severity_dv.add('E4:E100')
    
    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Closed"', allow_blank=True)
    sheet.add_data_validation(status_dv)
    status_dv.add('F4:F100')
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

def setup_dashboard_sheet(sheet, green_fill, red_fill, yellow_fill, border):
    """Setup the test progress dashboard sheet"""
    
    # Title
    sheet['A1'] = "AI SchoolOS - Test Progress Dashboard"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet.merge_cells('A1:F1')
    
    # Summary Statistics
    sheet['A3'] = "Summary Statistics"
    sheet['A3'].font = Font(size=14, bold=True)
    
    summary_data = [
        ['Total Tests', '=COUNTA(\'Detailed Test Cases\'!E:E)-1'],
        ['Passed', '=COUNTIF(\'Detailed Test Cases\'!E:E,"Pass")'],
        ['Failed', '=COUNTIF(\'Detailed Test Cases\'!E:E,"Fail")'],
        ['Blocked', '=COUNTIF(\'Detailed Test Cases\'!E:E,"Blocked")'],
        ['Pass Rate', '=IF(A4>0,B4/A4*100,0)']
    ]
    
    for row, (label, formula) in enumerate(summary_data, 4):
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=formula)
    
    # Portal Progress Table
    sheet['A8'] = "Portal Progress"
    sheet['A8'].font = Font(size=14, bold=True)
    
    portal_headers = ['Portal', 'Total Tests', 'Passed', 'Failed', 'Blocked', 'Pass Rate']
    for col, header in enumerate(portal_headers, 1):
        cell = sheet.cell(row=9, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = green_fill
        cell.border = border
    
    portal_data = [
        ['Landing Page', 6],
        ['Login Page', 8],
        ['Admin Portal', 9],
        ['Teacher Portal', 10],
        ['Student Portal', 10],
        ['Parent Portal', 10],
        ['Super Admin Portal', 10],
        ['Cross-Portal Access', 5],
        ['Backend Health', 3],
        ['TOTAL', '=SUM(B10:B18)']
    ]
    
    for row, (portal, total) in enumerate(portal_data, 10):
        sheet.cell(row=row, column=1, value=portal)
        sheet.cell(row=row, column=2, value=total)
        if portal == 'TOTAL':
            sheet.cell(row=row, column=1).font = Font(bold=True)
            sheet.cell(row=row, column=2).font = Font(bold=True)
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

def setup_reminders_sheet(sheet, border):
    """Setup the automated reminders sheet"""
    
    # Title
    sheet['A1'] = "AI SchoolOS - Automated Reminders"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet.merge_cells('A1:D1')
    
    # Headers
    headers = ['Reminder', 'Frequency', 'Description', 'Status']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        cell.border = border
    
    # Reminder data
    reminders_data = [
        ['Daily Progress Check', 'Daily', 'Review test progress and update status', '⬜'],
        ['Weekly Bug Review', 'Weekly', 'Review and prioritize issues', '⬜'],
        ['Portal Completion', 'On Portal Complete', 'Mark portal as fully tested', '⬜'],
        ['Final Review', 'On All Complete', 'Conduct final testing review', '⬜']
    ]
    
    for row, reminder in enumerate(reminders_data, 4):
        for col, value in enumerate(reminder, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = value
            cell.border = border
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

if __name__ == "__main__":
    try:
        filename = create_testing_checklist()
        print(f"\n🎉 Excel file created successfully!")
        print(f"📁 File location: {os.path.abspath(filename)}")
        print(f"📊 Features included:")
        print("   ✅ Conditional formatting for status")
        print("   ✅ Data validation for dropdowns")
        print("   ✅ Dashboard with progress tracking")
        print("   ✅ Issues tracker with screenshot column")
        print("   ✅ Automated progress calculation")
        print("   ✅ Comprehensive test cases")
        print("   ✅ Credentials reference")
        print("   ✅ Cross-portal access testing")
        print("   ✅ Role-based access control testing")
        print("\n🚀 Ready for systematic testing!")
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}") 